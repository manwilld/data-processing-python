"""Save TRS data to Excel using openpyxl."""

import os
import shutil
import numpy as np
from openpyxl import load_workbook, Workbook


def save_trs(run_name, axes, freq06_all, TRS06_all, seismic, output_dir):
    """Save TRS vs RRS data to Excel file.

    Parameters
    ----------
    run_name : str
    axes : list of str
    freq06_all : dict
        Keys like 'Table_X', values are freq06 arrays.
    TRS06_all : dict
        Keys like 'Table_X', values are TRS06 arrays.
    seismic : dict
        Contains RRS_h, RRS_v, freq72, lowCutoff, lowResonance.
    output_dir : str
    """
    # Column mapping: axis -> starting column (0-indexed)
    axis_to_col = {'X': 0, 'Y': 3, 'D': 6, 'Z': 9}

    output_file = os.path.join(output_dir, f'{run_name}_Table_TRSvsRRS.xlsx')

    # Try to copy template, or create new workbook
    template_path = os.path.join(os.path.dirname(os.path.dirname(__file__)),
                                 'templates', 'TRSvsRRS_Template.xlsx')
    if os.path.exists(template_path):
        shutil.copy2(template_path, output_file)
        wb = load_workbook(output_file)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # Create headers
        headers = ['Freq_X', 'RRS_X', 'TRS_X',
                   'Freq_Y', 'RRS_Y', 'TRS_Y',
                   'Freq_D', 'RRS_D', 'TRS_D',
                   'Freq_Z', 'RRS_Z', 'TRS_Z',
                   '', 'Low Resonance / Cutoff']
        for i, h in enumerate(headers):
            ws.cell(row=1, column=i + 1, value=h)
        # Row 2 is sub-headers
        sub_headers = ['Hz', 'g', 'g'] * 4 + ['', '']
        for i, h in enumerate(sub_headers):
            ws.cell(row=2, column=i + 1, value=h)

    freq72 = seismic['freq72']

    for axis in axes:
        table_col = f'Table_{axis}'
        if table_col not in freq06_all:
            continue

        current_freq06 = freq06_all[table_col]
        current_TRS06 = TRS06_all[table_col]

        # Get RRS at the 1/6 octave frequencies
        if axis in ('X', 'Y', 'D'):
            RRS_full = seismic['RRS_h']
        else:
            RRS_full = seismic['RRS_v']

        # Interpolate RRS at freq06 points
        current_RRS = np.interp(current_freq06, freq72, RRS_full)

        # Filter: only include frequencies > 1.0 Hz
        valid = current_freq06 > 1.0
        current_freq06 = current_freq06[valid]
        current_RRS = current_RRS[valid]
        current_TRS06 = current_TRS06[valid]

        # Round to 2 decimal places
        current_freq06 = np.round(current_freq06, 2)
        current_RRS = np.round(current_RRS, 2)
        current_TRS06 = np.round(current_TRS06, 2)

        # Write to Excel
        col_start = axis_to_col.get(axis, 0)
        row_start = 3  # Data starts at row 3

        for i in range(len(current_freq06)):
            ws.cell(row=row_start + i, column=col_start + 1,
                    value=float(current_freq06[i]))
            ws.cell(row=row_start + i, column=col_start + 2,
                    value=float(current_RRS[i]))
            ws.cell(row=row_start + i, column=col_start + 3,
                    value=float(current_TRS06[i]))

    # Write low resonance and cutoff
    ws.cell(row=3, column=14, value=seismic['lowResonance'])  # N3
    ws.cell(row=4, column=14, value=seismic['lowCutoff'])     # N4

    os.makedirs(output_dir, exist_ok=True)
    wb.save(output_file)
    print(f'TRS data saved to {output_file}')
