#!/usr/bin/env python3
"""Compare Python output vs MATLAB baseline.

Checks: trimmed CSV values/row counts, resonance CSV column order/values,
Excel column layout/row count/values, image pixel dimensions,
and per-plot mean pixel difference.
"""

import os
import sys
import numpy as np
import pandas as pd
from PIL import Image
from openpyxl import load_workbook

BASELINE_DIR = '/mnt/dropbox/_OpenClaw/data-processing-example/_matlab_baseline'
OUTPUT_DIR = '/mnt/dropbox/_OpenClaw/data-processing-example'
PIXEL_DIFF_THRESHOLD = 20.0


def check(name, passed, detail=''):
    status = 'PASS' if passed else 'FAIL'
    msg = f'  [{status}] {name}'
    if detail:
        msg += f' — {detail}'
    print(msg)
    return passed


def compare_csv(name, baseline_path, output_path, value_tol=0.01):
    """Compare two CSVs: column names, row count, values."""
    results = []
    if not os.path.exists(output_path):
        results.append(check(f'{name}: file exists', False, f'missing {output_path}'))
        return results

    df_b = pd.read_csv(baseline_path)
    df_o = pd.read_csv(output_path)

    # Row count
    results.append(check(
        f'{name}: row count',
        len(df_b) == len(df_o),
        f'baseline={len(df_b)}, output={len(df_o)}'
    ))

    # Column names and order
    cols_match = list(df_b.columns) == list(df_o.columns)
    results.append(check(
        f'{name}: column order',
        cols_match,
        f'baseline={list(df_b.columns)[:5]}..., output={list(df_o.columns)[:5]}...'
    ))

    # Value comparison (use overlapping columns and min row count)
    common_cols = [c for c in df_b.columns if c in df_o.columns]
    n_rows = min(len(df_b), len(df_o))
    if common_cols and n_rows > 0:
        b_vals = df_b[common_cols].iloc[:n_rows].values.astype(float)
        o_vals = df_o[common_cols].iloc[:n_rows].values.astype(float)
        max_diff = np.max(np.abs(b_vals - o_vals))
        results.append(check(
            f'{name}: values (tol={value_tol})',
            max_diff <= value_tol,
            f'max_diff={max_diff:.6f}'
        ))

    return results


def compare_excel(name, baseline_path, output_path):
    """Compare Excel files: column layout, row count, values."""
    results = []
    if not os.path.exists(output_path):
        results.append(check(f'{name}: file exists', False, f'missing {output_path}'))
        return results

    wb_b = load_workbook(baseline_path)
    wb_o = load_workbook(output_path)
    ws_b = wb_b.active
    ws_o = wb_o.active

    # Header row 1
    hdr_b = [ws_b.cell(row=1, column=c).value for c in range(1, 13)]
    hdr_o = [ws_o.cell(row=1, column=c).value for c in range(1, 13)]
    results.append(check(
        f'{name}: header row 1',
        hdr_b == hdr_o,
        f'baseline={hdr_b}, output={hdr_o}'
    ))

    # Header row 2
    hdr2_b = [ws_b.cell(row=2, column=c).value for c in range(1, 10)]
    hdr2_o = [ws_o.cell(row=2, column=c).value for c in range(1, 10)]
    results.append(check(
        f'{name}: header row 2',
        hdr2_b == hdr2_o,
        f'baseline={hdr2_b}, output={hdr2_o}'
    ))

    # Annotation cells
    for row, label in [(3, 'Lowest Resonance'), (4, 'Cuttoff Frequency')]:
        val_b = ws_b.cell(row=row, column=11).value
        val_o = ws_o.cell(row=row, column=11).value
        txt_b = ws_b.cell(row=row, column=12).value
        txt_o = ws_o.cell(row=row, column=12).value
        results.append(check(
            f'{name}: K{row}/L{row} ({label})',
            val_b == val_o and txt_b == txt_o,
            f'val: {val_b} vs {val_o}, txt: {txt_b!r} vs {txt_o!r}'
        ))

    # Data values (columns A-I, rows 3 onward)
    max_diff = 0
    row_count_b = 0
    row_count_o = 0
    for col in range(1, 10):
        for row in range(3, 200):
            vb = ws_b.cell(row=row, column=col).value
            vo = ws_o.cell(row=row, column=col).value
            if vb is not None:
                row_count_b += 1
            if vo is not None:
                row_count_o += 1
            if vb is not None and vo is not None:
                diff = abs(float(vb) - float(vo))
                max_diff = max(max_diff, diff)
    results.append(check(
        f'{name}: data cell count',
        row_count_b == row_count_o,
        f'baseline={row_count_b}, output={row_count_o}'
    ))
    results.append(check(
        f'{name}: data values (tol=0.01)',
        max_diff <= 0.01,
        f'max_diff={max_diff:.6f}'
    ))

    return results


def compare_images(name, baseline_dir, output_dir):
    """Compare image dimensions and pixel content."""
    results = []

    # Find all PNG files in baseline
    baseline_pngs = sorted([f for f in os.listdir(baseline_dir) if f.endswith('.png')])
    output_pngs = sorted([f for f in os.listdir(output_dir) if f.endswith('.png')])

    # Match by plot number prefix and plot type
    baseline_map = {}
    for f in baseline_pngs:
        baseline_map[f] = os.path.join(baseline_dir, f)

    output_map = {}
    for f in output_pngs:
        output_map[f] = os.path.join(output_dir, f)

    # Map by plot number (first token before _)
    def plot_key(fname):
        return fname.split('_')[0]

    b_by_key = {plot_key(f): f for f in baseline_pngs}
    o_by_key = {plot_key(f): f for f in output_pngs}

    results.append(check(
        f'{name}: plot count',
        len(baseline_pngs) == len(output_pngs),
        f'baseline={len(baseline_pngs)}, output={len(output_pngs)}'
    ))

    for key in sorted(b_by_key.keys(), key=lambda x: int(x)):
        if key not in o_by_key:
            results.append(check(f'{name}: plot {key} exists', False, 'missing'))
            continue

        b_path = os.path.join(baseline_dir, b_by_key[key])
        o_path = os.path.join(output_dir, o_by_key[key])

        b_img = Image.open(b_path)
        o_img = Image.open(o_path)

        # Dimension check
        dim_match = b_img.size == o_img.size
        results.append(check(
            f'{name}: plot {key} dimensions',
            dim_match,
            f'baseline={b_img.size}, output={o_img.size}'
        ))

        # Pixel difference (only if same size)
        if dim_match:
            b_arr = np.array(b_img).astype(float)
            o_arr = np.array(o_img).astype(float)
            # Handle RGBA vs RGB
            if b_arr.shape[-1] == 4 and o_arr.shape[-1] == 3:
                b_arr = b_arr[:, :, :3]
            elif b_arr.shape[-1] == 3 and o_arr.shape[-1] == 4:
                o_arr = o_arr[:, :, :3]
            elif b_arr.shape != o_arr.shape:
                b_arr = b_arr[:, :, :min(b_arr.shape[-1], o_arr.shape[-1])]
                o_arr = o_arr[:, :, :min(b_arr.shape[-1], o_arr.shape[-1])]

            mean_diff = np.mean(np.abs(b_arr - o_arr))
            results.append(check(
                f'{name}: plot {key} pixel diff',
                mean_diff <= PIXEL_DIFF_THRESHOLD,
                f'mean={mean_diff:.2f} (threshold={PIXEL_DIFF_THRESHOLD})'
            ))

    return results


def main():
    print('=' * 70)
    print('Python vs MATLAB Baseline Comparison')
    print('=' * 70)

    all_results = []

    # 1. Seismic trimmed CSV
    print('\n--- Seismic Trimmed CSV ---')
    all_results.extend(compare_csv(
        'Seismic CSV',
        os.path.join(BASELINE_DIR, 'Run_1_trimmed.csv'),
        os.path.join(OUTPUT_DIR, 'Run_1_trimmed.csv'),
    ))

    # 2. Resonance trimmed CSV
    print('\n--- Resonance Trimmed CSV ---')
    all_results.extend(compare_csv(
        'Resonance CSV',
        os.path.join(BASELINE_DIR, 'Run_1_resonance_trimmed.csv'),
        os.path.join(OUTPUT_DIR, 'Run_1_resonance_trimmed.csv'),
    ))

    # 3. Excel TRS vs RRS
    print('\n--- Excel TRS vs RRS ---')
    all_results.extend(compare_excel(
        'Excel',
        os.path.join(BASELINE_DIR, 'Run_1_Table_TRSvsRRS.xlsx'),
        os.path.join(OUTPUT_DIR, 'Run_1_Table_TRSvsRRS.xlsx'),
    ))

    # 4. Seismic plots
    print('\n--- Seismic Plots ---')
    all_results.extend(compare_images(
        'Seismic plots',
        os.path.join(BASELINE_DIR, 'Run_1 Plots_Seismic'),
        os.path.join(OUTPUT_DIR, 'Run_1 Plots_Seismic'),
    ))

    # 5. Resonance plots (UUT_1)
    print('\n--- Resonance Plots (UUT_1) ---')
    all_results.extend(compare_images(
        'UUT_1 resonance',
        os.path.join(BASELINE_DIR, 'UUT_1_Plots_Resonance'),
        os.path.join(OUTPUT_DIR, 'UUT_1_Plots_Resonance'),
    ))

    # 6. Resonance plots (UUT_2)
    print('\n--- Resonance Plots (UUT_2) ---')
    all_results.extend(compare_images(
        'UUT_2 resonance',
        os.path.join(BASELINE_DIR, 'UUT_2_Plots_Resonance'),
        os.path.join(OUTPUT_DIR, 'UUT_2_Plots_Resonance'),
    ))

    # Summary
    passed = sum(1 for r in all_results if r)
    failed = sum(1 for r in all_results if not r)
    total = len(all_results)

    print('\n' + '=' * 70)
    print(f'SUMMARY: {passed} passed, {failed} failed (out of {total} checks)')
    print('=' * 70)

    sys.exit(0 if failed == 0 else 1)


if __name__ == '__main__':
    main()
